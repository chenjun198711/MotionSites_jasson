#!/usr/bin/env python3
"""
Extract all prompt data from MotionSites-main project and output as Excel.
Sources:
  1. data/ms_prompts_with_text.json  (embedded prompt_text)
  2. data/catalog-meta.json           (full metadata for all entries)
  3. data/catalog-text-index.json     (maps id -> text file path)
  4. data/catalog-text/*.txt          (individual prompt text files)
  5. prompts/*.md                     (structured markdown prompts)
"""

import json
import os
import re
import sys

# Use openpyxl for rich formatting
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
PROMPTS_DIR = os.path.join(BASE_DIR, "prompts")
OUTPUT_PATH = os.path.join(BASE_DIR, "prompts_catalog.xlsx")


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def read_text(path):
    full = os.path.join(BASE_DIR, path)
    if not os.path.exists(full):
        return None
    with open(full, "r", encoding="utf-8") as f:
        return f.read().strip()


def parse_markdown_prompt(md_path):
    """Parse a markdown prompt file with YAML frontmatter."""
    with open(md_path, "r", encoding="utf-8") as f:
        content = f.read()

    # Extract frontmatter
    fm_match = re.match(r"^---\n(.*?)\n---\n(.*)", content, re.DOTALL)
    if not fm_match:
        return None

    fm_text = fm_match.group(1)
    body = fm_match.group(2)

    # Parse simple YAML frontmatter
    meta = {}
    for line in fm_text.split("\n"):
        m = re.match(r"^(\w+):\s*(.*)$", line)
        if m:
            key, val = m.group(1), m.group(2).strip()
            # Handle list values like [a, b, c]
            if val.startswith("[") and val.endswith("]"):
                val = val.strip("[]")
                items = [v.strip().strip('"').strip("'") for v in val.split(",") if v.strip()]
                meta[key] = items
            else:
                meta[key] = val.strip('"').strip("'")

    # Extract prompt from code block
    prompt_match = re.search(r"```text\n(.*?)\n```", body, re.DOTALL)
    prompt_text = prompt_match.group(1).strip() if prompt_match else ""

    # Extract description (line after first >)
    desc_match = re.search(r"^>\s*(.+)$", body, re.MULTILINE)
    description = desc_match.group(1).strip() if desc_match else ""

    # Extract source
    source_match = re.search(r"Origin:\s*(.+?)(?:\n|$)", body)
    source = source_match.group(1).strip() if source_match else ""

    # Extract license
    license_match = re.search(r"License:\s*(.+?)(?:\n|$)", body)
    license_val = license_match.group(1).strip() if license_match else ""

    return {
        "id": meta.get("id", ""),
        "title": meta.get("title", ""),
        "category": meta.get("category", ""),
        "type": meta.get("type", ""),
        "access": meta.get("access", ""),
        "tags": ", ".join(meta.get("tags", [])) if isinstance(meta.get("tags"), list) else meta.get("tags", ""),
        "palette": ", ".join(meta.get("palette", [])) if isinstance(meta.get("palette"), list) else meta.get("palette", ""),
        "created": meta.get("created", ""),
        "description": description,
        "prompt_text": prompt_text,
        "source": source,
        "license": license_val,
        "source_kind": "markdown",
    }


def main():
    # --- Source 1: ms_prompts_with_text.json (has prompt_text embedded) ---
    with_text_path = os.path.join(DATA_DIR, "ms_prompts_with_text.json")
    entries_with_text = load_json(with_text_path) if os.path.exists(with_text_path) else []
    print(f"Loaded {len(entries_with_text)} entries from ms_prompts_with_text.json")

    # Build lookup: id -> entry (with prompt_text)
    by_id = {}
    for entry in entries_with_text:
        eid = entry.get("id", "")
        if eid:
            by_id[eid] = {
                "id": eid,
                "title": entry.get("title", ""),
                "category": entry.get("category", ""),
                "type": entry.get("type", ""),
                "page_type": entry.get("page_type", ""),
                "is_free": entry.get("is_free"),
                "description": entry.get("description", ""),
                "prompt_text": entry.get("prompt_text", ""),
                "local_rel": entry.get("local_rel", ""),
                "local_kind": entry.get("local_kind", ""),
                "source_kind": entry.get("source_kind", ""),
                "source_id": entry.get("source_id"),
                "source_url": entry.get("source_url"),
                "source_path": entry.get("source_path"),
                "source_repo": entry.get("source_repo"),
                "source_license": entry.get("source_license"),
            }

    # --- Source 2: catalog-meta.json (all entries metadata, may lack prompt_text) ---
    meta_path = os.path.join(DATA_DIR, "catalog-meta.json")
    all_meta = load_json(meta_path) if os.path.exists(meta_path) else []
    print(f"Loaded {len(all_meta)} entries from catalog-meta.json")

    # --- Source 3: catalog-text-index.json (id -> text file path) ---
    text_index_path = os.path.join(DATA_DIR, "catalog-text-index.json")
    text_index = load_json(text_index_path) if os.path.exists(text_index_path) else {}
    print(f"Loaded {len(text_index)} text index entries")

    # --- Source 4: catalog-text.json (id -> text content) ---
    text_json_path = os.path.join(DATA_DIR, "catalog-text.json")
    text_json = load_json(text_json_path) if os.path.exists(text_json_path) else {}
    print(f"Loaded {len(text_json)} text entries from catalog-text.json")

    # Merge: add entries from catalog-meta that aren't already in by_id
    text_file_count = 0
    for entry in all_meta:
        eid = entry.get("id", "")
        if not eid:
            continue
        if eid in by_id:
            # Enrich existing entry with any missing fields
            existing = by_id[eid]
            for k, v in entry.items():
                if k not in existing or not existing.get(k):
                    existing[k] = v
            continue

        # New entry from catalog-meta - try to get prompt text
        prompt_text = ""

        # Try catalog-text.json first
        if eid in text_json:
            prompt_text = text_json[eid]
        # Then try text file
        elif eid in text_index:
            file_rel = text_index[eid]
            prompt_text = read_text(file_rel) or ""
            text_file_count += 1
        else:
            # Try reading data/catalog-text/{id}.txt directly
            direct_path = f"data/catalog-text/{eid}.txt"
            prompt_text = read_text(direct_path) or ""
            if prompt_text:
                text_file_count += 1

        by_id[eid] = {
            "id": eid,
            "title": entry.get("title", ""),
            "category": entry.get("category", ""),
            "type": entry.get("type", ""),
            "page_type": entry.get("page_type", ""),
            "is_free": entry.get("is_free"),
            "description": entry.get("description", ""),
            "prompt_text": prompt_text,
            "local_rel": entry.get("local_rel", ""),
            "local_kind": entry.get("local_kind", ""),
            "source_kind": entry.get("source_kind", ""),
            "image_preview_url": entry.get("image_preview_url", ""),
            "video_preview_url": entry.get("video_preview_url", ""),
            "created_at": entry.get("created_at", ""),
            "sort_order": entry.get("sort_order"),
            "has_text": entry.get("has_text", False),
            "text_len": entry.get("text_len", 0),
        }

    print(f"Read {text_file_count} text files from disk")

    # --- Source 5: prompts/*.md (structured markdown) ---
    md_count = 0
    if os.path.isdir(PROMPTS_DIR):
        for root, dirs, files in os.walk(PROMPTS_DIR):
            for fname in sorted(files):
                if fname.endswith(".md") and fname != "_TEMPLATE.md":
                    md_path = os.path.join(root, fname)
                    parsed = parse_markdown_prompt(md_path)
                    if parsed and parsed.get("id"):
                        eid = parsed["id"]
                        if eid not in by_id:
                            by_id[eid] = {
                                "id": eid,
                                "title": parsed["title"],
                                "category": parsed["category"],
                                "type": parsed["type"],
                                "access": parsed.get("access", ""),
                                "tags": parsed.get("tags", ""),
                                "palette": parsed.get("palette", ""),
                                "created": parsed.get("created", ""),
                                "description": parsed["description"],
                                "prompt_text": parsed["prompt_text"],
                                "source": parsed.get("source", ""),
                                "license": parsed.get("license", ""),
                                "source_kind": "markdown",
                            }
                            md_count += 1
                        else:
                            # Enrich existing with markdown data if prompt_text is empty
                            existing = by_id[eid]
                            if not existing.get("prompt_text") and parsed.get("prompt_text"):
                                existing["prompt_text"] = parsed["prompt_text"]
                            if not existing.get("tags") and parsed.get("tags"):
                                existing["tags"] = parsed.get("tags", "")
                            if not existing.get("palette") and parsed.get("palette"):
                                existing["palette"] = parsed.get("palette", "")
                            md_count += 1
    print(f"Processed {md_count} markdown prompt files")

    # --- Build final list, sorted by sort_order then title ---
    all_entries = list(by_id.values())
    # Sort: by sort_order if available, otherwise by title
    def sort_key(e):
        so = e.get("sort_order")
        if so is not None:
            return (0, so, e.get("title", ""))
        return (1, 0, e.get("title", ""))
    all_entries.sort(key=sort_key)

    print(f"\nTotal unique prompts: {len(all_entries)}")
    with_text = sum(1 for e in all_entries if e.get("prompt_text"))
    print(f"  With prompt text: {with_text}")
    print(f"  Without prompt text: {len(all_entries) - with_text}")

    # --- Generate Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "提示词目录"

    # Define styles
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="Microsoft YaHei", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    prompt_align = Alignment(vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Alternating row fill
    alt_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

    # Headers
    headers = [
        "序号",
        "ID",
        "提示词名称",
        "提示词描述",
        "分类",
        "类型",
        "页面类型",
        "是否免费",
        "标签",
        "配色方案",
        "完整提示词",
        "提示词长度",
        "来源类型",
        "来源仓库",
        "来源URL",
        "来源路径",
        "许可证",
        "创建时间",
        "预览文件路径",
        "预览类型",
        "图片预览URL",
        "视频预览URL",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    col_widths = {
        "A": 5,    # 序号
        "B": 28,   # ID
        "C": 24,   # 提示词名称
        "D": 45,   # 提示词描述
        "E": 16,   # 分类
        "F": 12,   # 类型
        "G": 12,   # 页面类型
        "H": 8,    # 是否免费
        "I": 30,   # 标签
        "J": 30,   # 配色方案
        "K": 80,   # 完整提示词
        "L": 10,   # 提示词长度
        "M": 14,   # 来源类型
        "N": 30,   # 来源仓库
        "O": 45,   # 来源URL
        "P": 40,   # 来源路径
        "Q": 14,   # 许可证
        "R": 22,   # 创建时间
        "S": 40,   # 预览文件路径
        "T": 10,   # 预览类型
        "U": 50,   # 图片预览URL
        "V": 50,   # 视频预览URL
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Write data rows
    for row_idx, entry in enumerate(all_entries, 2):
        is_free = entry.get("is_free")
        if is_free is True:
            free_str = "免费"
        elif is_free is False:
            free_str = "付费"
        else:
            free_str = entry.get("access", "")

        prompt_text = entry.get("prompt_text", "") or ""
        prompt_len = len(prompt_text)

        # Source fields
        source_kind = entry.get("source_kind", "") or ""
        source_repo = entry.get("source_repo", "") or ""
        source_url = entry.get("source_url", "") or ""
        source_path = entry.get("source_path", "") or ""
        source_license = entry.get("source_license", "") or entry.get("license", "") or ""
        source_origin = entry.get("source", "") or ""
        if source_origin and not source_repo:
            source_repo = source_origin

        created = entry.get("created_at", "") or entry.get("created", "") or ""

        row_data = [
            row_idx - 1,                                          # 序号
            entry.get("id", ""),                                  # ID
            entry.get("title", ""),                               # 提示词名称
            entry.get("description", ""),                         # 提示词描述
            entry.get("category", ""),                            # 分类
            entry.get("type", ""),                                # 类型
            entry.get("page_type", ""),                           # 页面类型
            free_str,                                             # 是否免费
            entry.get("tags", ""),                                # 标签
            entry.get("palette", ""),                             # 配色方案
            prompt_text,                                          # 完整提示词
            prompt_len if prompt_len > 0 else "",                # 提示词长度
            source_kind,                                          # 来源类型
            source_repo,                                          # 来源仓库
            source_url,                                           # 来源URL
            source_path,                                          # 来源路径
            source_license,                                       # 许可证
            created,                                              # 创建时间
            entry.get("local_rel", ""),                           # 预览文件路径
            entry.get("local_kind", ""),                          # 预览类型
            entry.get("image_preview_url", ""),                   # 图片预览URL
            entry.get("video_preview_url", ""),                   # 视频预览URL
        ]

        for col_idx, value in enumerate(row_data, 1):
            # Convert lists/dicts to string representation
            if isinstance(value, (list, dict)):
                value = ", ".join(str(v) for v in value) if isinstance(value, list) else str(value)
            elif value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            if col_idx == 11:  # 完整提示词列
                cell.alignment = prompt_align
            else:
                cell.alignment = cell_align
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto filter
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(all_entries) + 1}"

    # --- Create a summary sheet ---
    ws2 = wb.create_sheet("统计摘要")

    # Count by category
    category_counts = {}
    type_counts = {}
    free_counts = {"免费": 0, "付费": 0, "未知": 0}
    source_counts = {}
    with_text_count = 0
    without_text_count = 0

    for entry in all_entries:
        cat = entry.get("category", "未分类")
        category_counts[cat] = category_counts.get(cat, 0) + 1

        t = entry.get("type", "未知")
        type_counts[t] = type_counts.get(t, 0) + 1

        is_free = entry.get("is_free")
        if is_free is True:
            free_counts["免费"] += 1
        elif is_free is False:
            free_counts["付费"] += 1
        else:
            free_counts["未知"] += 1

        sk = entry.get("source_kind", "未知")
        source_counts[sk] = source_counts.get(sk, 0) + 1

        if entry.get("prompt_text"):
            with_text_count += 1
        else:
            without_text_count += 1

    # Summary sheet styles
    title_font = Font(name="Microsoft YaHei", size=14, bold=True, color="2B5797")
    section_font = Font(name="Microsoft YaHei", size=12, bold=True, color="2B5797")
    data_font = Font(name="Microsoft YaHei", size=10)
    label_font = Font(name="Microsoft YaHei", size=10, bold=True)

    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 4
    ws2.column_dimensions["E"].width = 30
    ws2.column_dimensions["F"].width = 12

    row = 1
    ws2.cell(row=row, column=2, value="提示词统计摘要").font = title_font
    row += 2

    # Overview
    ws2.cell(row=row, column=2, value="总览").font = section_font
    row += 1
    overview = [
        ("提示词总数", len(all_entries)),
        ("有完整提示词文本", with_text_count),
        ("无提示词文本", without_text_count),
        ("提示词文本覆盖率", f"{with_text_count / len(all_entries) * 100:.1f}%"),
    ]
    for label, val in overview:
        ws2.cell(row=row, column=2, value=label).font = label_font
        ws2.cell(row=row, column=3, value=val).font = data_font
        row += 1

    row += 1
    # Category distribution
    ws2.cell(row=row, column=2, value="分类分布").font = section_font
    row += 1
    ws2.cell(row=row, column=2, value="分类").font = header_font
    ws2.cell(row=row, column=3, value="数量").font = header_font
    ws2.cell(row=row, column=2).fill = header_fill
    ws2.cell(row=row, column=3).fill = header_fill
    ws2.cell(row=row, column=2).alignment = header_align
    ws2.cell(row=row, column=3).alignment = header_align
    row += 1
    for cat, count in sorted(category_counts.items(), key=lambda x: -x[1]):
        ws2.cell(row=row, column=2, value=cat).font = data_font
        ws2.cell(row=row, column=3, value=count).font = data_font
        row += 1

    row += 1
    # Type distribution
    ws2.cell(row=row, column=2, value="类型分布").font = section_font
    row += 1
    ws2.cell(row=row, column=2, value="类型").font = header_font
    ws2.cell(row=row, column=3, value="数量").font = header_font
    ws2.cell(row=row, column=2).fill = header_fill
    ws2.cell(row=row, column=3).fill = header_fill
    ws2.cell(row=row, column=2).alignment = header_align
    ws2.cell(row=row, column=3).alignment = header_align
    row += 1
    for t, count in sorted(type_counts.items(), key=lambda x: -x[1]):
        ws2.cell(row=row, column=2, value=t).font = data_font
        ws2.cell(row=row, column=3, value=count).font = data_font
        row += 1

    row += 1
    # Free/Paid
    ws2.cell(row=row, column=2, value="免费/付费分布").font = section_font
    row += 1
    ws2.cell(row=row, column=2, value="类型").font = header_font
    ws2.cell(row=row, column=3, value="数量").font = header_font
    ws2.cell(row=row, column=2).fill = header_fill
    ws2.cell(row=row, column=3).fill = header_fill
    ws2.cell(row=row, column=2).alignment = header_align
    ws2.cell(row=row, column=3).alignment = header_align
    row += 1
    for k, v in free_counts.items():
        ws2.cell(row=row, column=2, value=k).font = data_font
        ws2.cell(row=row, column=3, value=v).font = data_font
        row += 1

    row += 1
    # Source distribution
    ws2.cell(row=row, column=2, value="来源分布").font = section_font
    row += 1
    ws2.cell(row=row, column=2, value="来源").font = header_font
    ws2.cell(row=row, column=3, value="数量").font = header_font
    ws2.cell(row=row, column=2).fill = header_fill
    ws2.cell(row=row, column=3).fill = header_fill
    ws2.cell(row=row, column=2).alignment = header_align
    ws2.cell(row=row, column=3).alignment = header_align
    row += 1
    for s, count in sorted(source_counts.items(), key=lambda x: -x[1]):
        ws2.cell(row=row, column=2, value=s).font = data_font
        ws2.cell(row=row, column=3, value=count).font = data_font
        row += 1

    # Save
    wb.save(OUTPUT_PATH)
    print(f"\nExcel saved to: {OUTPUT_PATH}")
    print(f"  Sheet 1 '提示词目录': {len(all_entries)} rows")
    print(f"  Sheet 2 '统计摘要': summary stats")


if __name__ == "__main__":
    main()
